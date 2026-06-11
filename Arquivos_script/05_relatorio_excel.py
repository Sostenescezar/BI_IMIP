"""
SAD IMIP - Marketing Digital
Script 05: Relatório Consolidado + Exportação Excel
Gera um arquivo .xlsx com múltiplas abas prontas para o SAD:
  - Resumo Executivo
  - Segmentação de Pacientes (RFM)
  - Perfil dos Médicos
  - Investidores por Tier
  - Engajamento por Rede
  - ROI por Público
  - Funil de Conversão
  - Recomendações

Depende de: imip_sad.db (scripts 01–04 já executados)
"""

import sqlite3
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import os

DB_PATH     = "./imip_sad.db"
OUTPUT_DIR  = "./relatorios/"
DATA_HOJE   = datetime.now().strftime("%Y%m%d")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"SAD_IMIP_Marketing_{DATA_HOJE}.xlsx")

os.makedirs(OUTPUT_DIR, exist_ok=True)

AZUL_IMIP    = "003087"   # azul institucional (ajuste conforme identidade do IMIP)
VERDE_IMIP   = "006B3C"
AMARELO      = "FFC300"
CINZA_CLARO  = "F2F2F2"
BRANCO       = "FFFFFF"

def conectar():
    return sqlite3.connect(DB_PATH)

def ler_tabela(nome, fallback=None):
    try:
        conn = conectar()
        df = pd.read_sql(f"SELECT * FROM {nome}", conn)
        conn.close()
        return df
    except Exception:
        return fallback if fallback is not None else pd.DataFrame()


# ─────────────────────────────────────────────
# HELPERS DE FORMATAÇÃO EXCEL
# ─────────────────────────────────────────────
def estilo_cabecalho(ws, row, cor_hex=AZUL_IMIP):
    fill = PatternFill("solid", fgColor=cor_hex)
    font = Font(bold=True, color=BRANCO, size=11)
    aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    brd  = Border(
        bottom=Side(style="medium", color=BRANCO),
        right=Side(style="thin",   color=BRANCO)
    )
    for cell in ws[row]:
        cell.fill, cell.font, cell.alignment, cell.border = fill, font, aln, brd

def auto_largura(ws, min_w=12, max_w=45):
    for col in ws.columns:
        comprimento = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max(comprimento + 2, min_w), max_w)

def zebra(ws, inicio_linha=2, cor_par=CINZA_CLARO):
    fill_par = PatternFill("solid", fgColor=cor_par)
    for i, row in enumerate(ws.iter_rows(min_row=inicio_linha)):
        if i % 2 == 0:
            for cell in row:
                cell.fill = fill_par

def escrever_df(ws, df, titulo=None, cor_cab=AZUL_IMIP, linha_inicio=1):
    """Escreve um DataFrame na aba com título, cabeçalho colorido e zebra."""
    lin = linha_inicio
    if titulo:
        ws.cell(lin, 1, titulo).font = Font(bold=True, size=13, color=AZUL_IMIP)
        lin += 1
    # cabeçalho
    for j, col in enumerate(df.columns, 1):
        ws.cell(lin, j, str(col))
    estilo_cabecalho(ws, lin, cor_cab)
    lin_cab = lin
    lin += 1
    # dados
    for _, row in df.iterrows():
        for j, val in enumerate(row, 1):
            cell = ws.cell(lin, j)
            cell.value = val if not (isinstance(val, float) and np.isnan(val)) else None
            cell.alignment = Alignment(horizontal="center")
        lin += 1
    zebra(ws, lin_cab + 1)
    auto_largura(ws)
    return lin


# ─────────────────────────────────────────────
# RESUMO EXECUTIVO (calculado na hora)
# ─────────────────────────────────────────────
def gerar_resumo_executivo():
    conn = conectar()
    linhas = []

    def sql_val(q, default="N/D"):
        try:
            r = pd.read_sql(q, conn)
            v = r.iloc[0, 0]
            return v if v is not None else default
        except Exception:
            return default

    linhas += [
        ("PÚBLICO",              "MÉTRICA",                        "VALOR"),
        ("Pacientes",            "Total de registros",             sql_val("SELECT COUNT(*) FROM pacientes")),
        ("Pacientes",            "Segmento Fiel",                  sql_val("SELECT COUNT(*) FROM pacientes_rfm WHERE segmento_rfm='Paciente Fiel'")),
        ("Pacientes",            "Segmento Inativo",               sql_val("SELECT COUNT(*) FROM pacientes_rfm WHERE segmento_rfm='Inativo'")),
        ("Médicos",              "Total de registros",             sql_val("SELECT COUNT(*) FROM medicos")),
        ("Médicos",              "Residentes",                     sql_val("SELECT COUNT(*) FROM medicos WHERE tipo_medico='Residente'")),
        ("Médicos",              "Staff",                          sql_val("SELECT COUNT(*) FROM medicos WHERE tipo_medico='Staff'")),
        ("Investidores",         "Total de registros",             sql_val("SELECT COUNT(*) FROM investidores")),
        ("Investidores",         "Tier Estratégico",               sql_val("SELECT COUNT(*) FROM investidores_segmentados WHERE tier_comunicacao='Estratégico'")),
        ("Campanhas",            "Investimento total (R$)",        sql_val("SELECT ROUND(SUM(custo_total),2) FROM campanhas")),
        ("Campanhas",            "Receita gerada (R$)",            sql_val("SELECT ROUND(SUM(receita_gerada),2) FROM campanhas")),
        ("Campanhas",            "ROAS médio",                     sql_val("SELECT ROUND(AVG(roas),2) FROM performance_campanhas")),
        ("Campanhas",            "Melhor canal (conversões)",      sql_val("SELECT canal FROM atribuicao_canais ORDER BY conversoes DESC LIMIT 1")),
        ("Redes Sociais",        "Plataforma mais engajada",       sql_val("SELECT rede_social FROM metricas_por_rede ORDER BY engajamento_medio_pct DESC LIMIT 1")),
        ("Redes Sociais",        "Melhor engajamento médio (%)",   sql_val("SELECT MAX(engajamento_medio_pct) FROM metricas_por_rede")),
    ]
    conn.close()
    df = pd.DataFrame(linhas[1:], columns=linhas[0])
    return df


# ─────────────────────────────────────────────
# ABA DE RECOMENDAÇÕES
# ─────────────────────────────────────────────
def gerar_recomendacoes():
    dados = [
        ("Público",      "Área",          "Recomendação",                                                      "Prioridade"),
        ("Pacientes",    "CRM Digital",   "Ativar campanha de reengajamento para segmento 'Inativo' via e-mail/WhatsApp", "Alta"),
        ("Pacientes",    "Conteúdo",      "Criar conteúdo educativo de saúde preventiva para público 'Adulto' e 'Idoso'",  "Média"),
        ("Pacientes",    "SEO/Ads",       "Investir em palavras-chave locais (Recife, PE) para captação orgânica",         "Alta"),
        ("Médicos Res.", "LinkedIn",      "Publicar posts de capacitação e residência médica para atrair novos residentes","Alta"),
        ("Médicos Res.", "Conteúdo",      "Série de conteúdo sobre oportunidades de carreira no IMIP",                   "Média"),
        ("Médicos Staff","LinkedIn",      "Destacar casos clínicos e publicações científicas para reputação institucional","Alta"),
        ("Médicos Staff","E-mail",        "Newsletter mensal com novidades do hospital, eventos e pesquisas",              "Média"),
        ("Investidores", "Relatórios",    "Enviar relatório trimestral de impacto social e financeiro para tier Estratégico","Alta"),
        ("Investidores", "Eventos",       "Convidar tier Prioritário para eventos e apresentações institucionais",         "Média"),
        ("Geral",        "Redes Sociais", "Concentrar postagens nos melhores horários identificados na análise",           "Alta"),
        ("Geral",        "Budget",        "Redirecionar verba dos canais deficitários para os canais de alto ROAS",        "Alta"),
        ("Geral",        "A/B Tests",     "Testar 2 variações criativas por público a cada 30 dias",                      "Média"),
    ]
    df = pd.DataFrame(dados[1:], columns=dados[0])
    return df


# ─────────────────────────────────────────────
# MAIN — MONTAGEM DO EXCEL
# ─────────────────────────────────────────────
def gerar_relatorio():
    print(f"\nGerando relatório: {OUTPUT_FILE}")

    abas = {
        "Resumo Executivo":      (gerar_resumo_executivo(),           AZUL_IMIP),
        "Pacientes RFM":         (ler_tabela("pacientes_rfm"),        AZUL_IMIP),
        "Médicos Clusters":      (ler_tabela("medicos_clusters"),     VERDE_IMIP),
        "Investidores":          (ler_tabela("investidores_segmentados"), VERDE_IMIP),
        "Engajamento Redes":     (ler_tabela("metricas_por_rede"),    AZUL_IMIP),
        "Top Posts":             (ler_tabela("top_posts"),            AZUL_IMIP),
        "ROI por Público":       (ler_tabela("roi_por_publico_canal"),VERDE_IMIP),
        "Funil de Conversão":    (ler_tabela("funil_conversao"),      VERDE_IMIP),
        "Atribuição Canais":     (ler_tabela("atribuicao_canais"),    AZUL_IMIP),
        "Recomendações":         (gerar_recomendacoes(),              AMARELO),
    }

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for nome_aba, (df, cor) in abas.items():
            if df.empty:
                df = pd.DataFrame([["Dados não disponíveis — execute os scripts 01 a 04 primeiro."]])
            df.to_excel(writer, sheet_name=nome_aba[:31], index=False, startrow=1)

    # Pós-formatação com openpyxl
    wb = load_workbook(OUTPUT_FILE)
    for nome_aba, (df, cor) in abas.items():
        nome = nome_aba[:31]
        if nome not in wb.sheetnames:
            continue
        ws = wb[nome]
        estilo_cabecalho(ws, 2, cor)
        zebra(ws, 3)
        auto_largura(ws)
        # Título da aba
        ws["A1"] = f"SAD IMIP — {nome_aba}   |   Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A1"].font = Font(bold=True, size=11, color=AZUL_IMIP)

    # Adiciona gráfico na aba ROI
    if "ROI por Público" in wb.sheetnames:
        ws_roi = wb["ROI por Público"]
        try:
            chart = BarChart()
            chart.type   = "col"
            chart.title  = "ROI (%) por Público e Canal"
            chart.y_axis.title = "ROI %"
            chart.x_axis.title = "Campanha"
            n_linhas = ws_roi.max_row
            data = Reference(ws_roi, min_col=5, max_col=5, min_row=2, max_row=n_linhas)
            cats = Reference(ws_roi, min_col=1, max_col=2, min_row=3, max_row=n_linhas)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            ws_roi.add_chart(chart, "K2")
        except Exception:
            pass

    wb.save(OUTPUT_FILE)
    print(f"✓ Relatório salvo em: {OUTPUT_FILE}")
    return OUTPUT_FILE


if __name__ == "__main__":
    print("=== GERAÇÃO DE RELATÓRIO - SAD IMIP ===")
    gerar_relatorio()
    print("=== CONCLUÍDO ===")
